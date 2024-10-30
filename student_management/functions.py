# functions.py
from tkinter import messagebox

from student_management.data import SAMPLE_DATA, SAMPLE_ABSENCES, delete_student_from_db
from student_management.getter_data import get_sample_data, get_sample_absences


def search_student_in_sample(query):
    """Searches for students in the sample data based on name or MSSV."""
    results = [s for s in SAMPLE_DATA if query.lower() in s['name'].lower() or s['mssv'] == query]
    return results

# Hàm cập nhật TreeView
def update_treeview(data, tree):
    """Cập nhật Treeview với dữ liệu mới."""
    # Xóa dữ liệu cũ trong TreeView
    for item in tree.get_children():
        tree.delete(item)

    # Duyệt qua từng sinh viên để thêm thông tin
    for student in data:
        mssv = student['mssv']
        ma_lop = student['ma_lop']

        # Tính số tiết vắng
        total_absences = sum(absence['tong_tiet'] for absence in get_sample_absences() if
                             absence['mssv'] == mssv and absence['ma_lop'] == ma_lop)

        # Tính tỷ lệ vắng
        total_classes = 30
        absence_text = f"{total_absences}/{total_classes}" if total_classes > 0 else f"{total_absences}/0"

        # Thêm dữ liệu vào TreeView
        tree.insert("", "end", values=(
            student['dot'], student['ma_lop'], student['ten_mh'], student['name'], mssv, absence_text))

def search_student(query, tree):
    """Search function for students in the sample data."""
    results = search_student_in_sample(query)
    update_treeview(results, tree)


import tkinter.messagebox as messagebox


def delete_student(tree):
    selected_item = tree.selection()
    if selected_item:
        item = tree.item(selected_item)
        values = item['values']
        mssv = str(values[4])
        ma_lop = str(values[1]).zfill(12)  # Điều chỉnh độ dài cho phù hợp với định dạng trong cơ sở dữ liệu

        # Hiển thị hộp thoại xác nhận
        confirm = messagebox.askyesno("Xác nhận xóa",
                                      f"Bạn có chắc chắn muốn xóa sinh viên:\nMSSV: {mssv}\nMã lớp: {ma_lop}?")

        if confirm:  # Nếu người dùng chọn "Yes"
            print(f"Deleting student: MSSV: {mssv}, Ma lop: {ma_lop}")  # Dòng để kiểm tra

            # Xóa sinh viên khỏi cơ sở dữ liệu dựa trên MSSV và ma_lop
            delete_student_from_db(mssv, ma_lop)

            # Làm mới SAMPLE_DATA với dữ liệu cơ sở dữ liệu đã cập nhật
            global SAMPLE_DATA
            SAMPLE_DATA = get_sample_data()

            # Làm mới TreeView với dữ liệu đã cập nhật
            update_treeview(SAMPLE_DATA, tree)
    else:
        messagebox.showwarning("No selection", "Please select a student to delete.")


def on_tree_select(event, tree):
    """Handles selecting an item in the student Treeview."""
    selected_item = tree.selection()
    if selected_item:
        item = tree.item(selected_item)
        values = item['values']
        return values  # Return student data
    return None


def sort_students_in_treeview(tree, criteria):
    """Sorts the students in the Treeview based on the selected criteria."""
    if criteria == 'total_absences':
        # Tính tổng số buổi vắng cho từng sinh viên
        absences_count = {}
        for absence in SAMPLE_ABSENCES:
            key = (absence['mssv'], absence['ma_lop'])
            absences_count[key] = absences_count.get(key, 0) + absence['vang_phep'] + absence['vang_khong_phep']

        # Kết hợp với dữ liệu sinh viên
        sorted_data = sorted(SAMPLE_DATA, key=lambda s: absences_count.get((s['mssv'], s['ma_lop']), 0), reverse=True)
    elif criteria == 'name':
        sorted_data = sorted(SAMPLE_DATA, key=lambda x: x['name'])
    elif criteria == 'class':
        sorted_data = sorted(SAMPLE_DATA, key=lambda x: x['ma_lop'])
    elif criteria == 'subject':
        sorted_data = sorted(SAMPLE_DATA, key=lambda x: x['ten_mh'])
    else:
        # Trường hợp mặc định: không sắp xếp, giữ nguyên thứ tự ban đầu
        sorted_data = SAMPLE_DATA
    update_treeview(sorted_data, tree)


def filter_by_class(tree, class_name):
    """Lọc sinh viên theo lớp và hiển thị kết quả trong Treeview."""
    # Lọc dữ liệu sinh viên theo mã lớp
    filtered_data = [sinh_vien for sinh_vien in SAMPLE_DATA if sinh_vien['ma_lop'] == class_name]
    update_treeview(filtered_data, tree)


def filter_by_absences(tree, min_absences):
    """Lọc sinh viên có số buổi vắng lớn hơn giá trị min_absences và hiển thị kết quả trong Treeview."""
    # Tính tổng số buổi vắng cho từng sinh viên
    absences_count = {}
    for absence in SAMPLE_ABSENCES:
        key = (absence['mssv'], absence['ma_lop'])
        absences_count[key] = absences_count.get(key, 0) + absence['vang_phep'] + absence['vang_khong_phep']

    # Lọc dữ liệu sinh viên dựa trên số buổi vắng
    filtered_data = [sinh_vien for sinh_vien in SAMPLE_DATA
                     if absences_count.get((sinh_vien['mssv'], sinh_vien['ma_lop']), 0) > min_absences]
    update_treeview(filtered_data, tree)


def apply_filters_and_sort(tree, selected_class=None, min_absences=None, sort_criteria=None, ascending=True):
    """Áp dụng lọc và sắp xếp cho dữ liệu sinh viên và hiển thị kết quả trong Treeview."""
    # Lọc dữ liệu
    filtered_data = filter_data(selected_class, min_absences)
    # Sắp xếp dữ liệu đã lọc
    sorted_data = sort_data(filtered_data, sort_criteria, ascending=ascending)
    # Cập nhật lại Treeview với dữ liệu đã lọc và sắp xếp
    update_treeview(sorted_data, tree)


def filter_data(selected_class, min_absences):
    """Trả về danh sách sinh viên đã lọc theo lớp và số buổi vắng."""
    # Tính tổng số buổi vắng cho từng sinh viên
    absences_count = {}
    for absence in SAMPLE_ABSENCES:
        key = (absence['mssv'], absence['ma_lop'])
        absences_count[key] = absences_count.get(key, 0) + absence['vang_phep'] + absence['vang_khong_phep']

    # Lọc dữ liệu
    filtered_data = get_sample_data()
    if selected_class:
        filtered_data = [sv for sv in filtered_data if sv['ma_lop'] == selected_class]
    if min_absences is not None:
        filtered_data = [sv for sv in filtered_data if absences_count.get((sv['mssv'], sv['ma_lop']), 0) > min_absences]

    return filtered_data


def sort_data(data, criteria, ascending=True):
    """Sắp xếp dữ liệu sinh viên dựa trên tiêu chí đã chọn và thứ tự sắp xếp."""
    if criteria == 'total_absences':
        # Sắp xếp theo số buổi vắng
        absences_count = {(absence['mssv'], absence['ma_lop']): 0 for absence in get_sample_absences()}
        for absence in get_sample_absences():
            key = (absence['mssv'], absence['ma_lop'])
            absences_count[key] += absence['vang_phep'] + absence['vang_khong_phep']

        return sorted(data, key=lambda s: absences_count.get((s['mssv'], s['ma_lop']), 0), reverse=not ascending)
    elif criteria == 'name':
        return sorted(data, key=lambda x: x['name'], reverse=not ascending)
    elif criteria == 'class':
        return sorted(data, key=lambda x: x['ma_lop'], reverse=not ascending)
    elif criteria == 'subject':
        return sorted(data, key=lambda x: x['ten_mh'], reverse=not ascending)
    else:
        # Không sắp xếp nếu không có tiêu chí
        return data


import sqlite3
import openpyxl


def load_students_from_excel(filename):
    # Mở file Excel
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Lấy thông tin lớp học từ các ô cụ thể
    dot = sheet['C6'].value
    ma_lop = sheet['C8'].value
    ten_mh = sheet['C9'].value

    # Kết nối tới database
    with sqlite3.connect("student_data.db") as conn:
        cursor = conn.cursor()

        # Duyệt qua các dòng chứa dữ liệu sinh viên
        for row in range(14, sheet.max_row + 1):
            name = f"{sheet[f'C{row}'].value} {sheet[f'D{row}'].value}"  # Ghép họ tên
            mssv = sheet[f'B{row}'].value  # MSSV nằm ở cột B
            if name and mssv:
                # Chèn dữ liệu vào bảng students (bỏ qua nếu sinh viên đã tồn tại)
                cursor.execute("""INSERT OR IGNORE INTO students (dot, ma_lop, ten_mh, name, mssv) 
                                  VALUES (?, ?, ?, ?, ?)""",
                               (dot, ma_lop, ten_mh, name, mssv))

                # Duyệt qua các cột chứa thông tin điểm danh, bắt đầu từ cột G (cột 7)
                for col in range(7, 25, 3):
                    ngay = sheet.cell(row=12, column=col).value  # Lấy giá trị ngày từ hàng 12
                    vang_phep = 0
                    vang_khong_phep = 0
                    if sheet.cell(row=row, column=col).value == 'P':  # Cột chứa ký hiệu P (vắng phép)
                        vang_phep = sheet.cell(row=row, column=col + 1).value or 0  # Số tiết vắng phép
                    elif sheet.cell(row=row, column=col).value == 'K':  # Cột chứa ký hiệu K (vắng không phép)
                        vang_khong_phep = sheet.cell(row=row, column=col + 1).value or 0  # Số tiết vắng không phép
                    tong_tiet = vang_phep + vang_khong_phep

                    if ngay:  # Nếu có ngày thì lưu vào bảng absences
                        cursor.execute("""INSERT INTO absences (mssv, ma_lop, ngay, vang_phep, vang_khong_phep, tong_tiet) 
                                          VALUES (?, ?, ?, ?, ?, ?)""",
                                       (mssv, ma_lop, ngay, vang_phep, vang_khong_phep, tong_tiet))

        # Lưu các thay đổi vào cơ sở dữ liệu
        conn.commit()

    print(f"Đã load dữ liệu từ file {filename} thành công.")
